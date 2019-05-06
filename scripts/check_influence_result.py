# -*- coding: utf-8 -*-
"""
Created on Tue Mar 19 12:29:43 2019

@author: Administrator
"""

import numpy as np
from openpyxl import Workbook
import os
#%%
BASE_DIR = 'output6/'
RESULT_STR = 'rbf_carair_results'
DATA_DIR = 'C:/Tang/data/car_airplane/'

#%%
# 原始数据集
orig_dataset = np.load('C:/Tang/data/car_airplane/dataset_car-airplane_train-1000_test-300.npz')
Y_train=orig_dataset['Y_train']

#%%
index_titles = {0:'label',
                1:'postive influence count',
                2:'postive influence sum',
                3:'negative influence count',
                4:'negative influence sum',
                5:'top5 count',
                6:'top5 influence sum',
                7:'last5 count',
                8:'last5 influence sum',
                9:'influence sum',
                10:'postive influence count same cata',
                11:'postive influence count diff cata',
                12:'negative influence count same cata',
                13:'negative influence count diff cata',
                14:'top5 count same cata',
                15:'top5 count diff cata',
                16:'last5 count same cata',
                17:'last5 count diff cata'}

#%%
# 创建出现在top5中过的训练点的训练集
top5_Xtrain = np.zeros((80, 299, 299, 3))
top5_Ytrain = np.zeros((80))


for i, train_idx in enumerate(np.argsort(top5count[:, 0])[-78:]):
    top5_Xtrain[i] = X_train[train_idx]
    top5_Ytrain[i] = Y_train[train_idx]

# 额外增加的两个飞机训练点
top5_Xtrain[78] = X_train[213]
top5_Ytrain[78] = Y_train[213]
top5_Xtrain[79] = X_train[1382]
top5_Ytrain[79] = Y_train[1382]

classes = ['car', 'airplane']
data_filename = os.path.join(DATA_DIR, 'dataset_%s_train-%s_test-%s.npz' % ('-'.join(classes), 40, 300))
np.savez_compressed(data_filename, X_train=top5_Xtrain, Y_train=top5_Ytrain, X_test=X_test, Y_test=Y_test)

#%%
# 把top5,last5,正向影响计数，负向影响计数等都存进一个文件
# =============================================================================
# index:index
# 0:label
# 1:postive influence count
# 2:postive influence sum
# 3:negative influence count
# 4:negative influence sum
# 5:top5 count
# 6:top5 influence sum
# 7:last5 count
# 8:last5 influence sum
# 9:influence sum
# =============================================================================
# index不用单独一列存
infl_res = np.zeros((2000, 10))
for i in range(600):
    t = np.load(BASE_DIR + RESULT_STR + '_%s.npz' % i)
    infl = t['inception_predicted_loss_diffs']
    for idx in np.argsort(infl)[-50:]:
        infl_res[idx][5] += 1
        infl_res[idx][6] += infl[idx]
    for idx in np.argsort(infl)[:50]:
        infl_res[idx][7] += 1
        infl_res[idx][8] += infl[idx]
    for j in range(2000):
        infl_res[j][0] = Y_train[j]
        infl_res[j][9] += infl[j]
        if infl[j]>=0:
            infl_res[j][1] += 1
            infl_res[j][2] += infl[j]
        else:
            infl_res[j][3] += 1
            infl_res[j][4] += infl[j]

        
np.savez(BASE_DIR + 'inception_influence_result_top50', result = infl_res)

wb = Workbook()
ws = wb.active
for i in range(2000):
    ws.cell(row = i+1, column = 1, value = i)
    for j in range(10):
        ws.cell(row = i+1, column = j+2, value = infl_res[i][j])

wb.save(BASE_DIR + "inception_influence_result_carair_top50.xlsx")
#%%
# 观察一下dog_fish数据集的情况
BASE_DIR = 'output/'
RESULT_STR = 'rbf_dogfish_results'
DATA_DIR = 'C:/Tang/data/dog_fish/'

orig_dataset = np.load(DATA_DIR + 'dataset_dog-fish_train-900_test-300.npz')
Y_train=orig_dataset['Y_train']

infl_res = np.zeros((1800, 18))
for i in range(600):
    t = np.load(BASE_DIR + RESULT_STR + '_%s.npz' % i)
    infl = t['inception_predicted_loss_diffs']
    for idx in np.argsort(infl)[-5:]:
        infl_res[idx][5] += 1
        infl_res[idx][6] += infl[idx]
        if t['flipped_idx'][0]:
            infl_res[idx][14] += 1
        else:
            infl_res[idx][15] += 1
    for idx in np.argsort(infl)[:5]:
        infl_res[idx][7] += 1
        infl_res[idx][8] += infl[idx]
        if t['flipped_idx'][0]:
            infl_res[idx][16] += 1
        else:
            infl_res[idx][17] += 1
    for j in range(1800):
        infl_res[j][0] = Y_train[j]
        infl_res[j][9] += infl[j]
        if infl[j]>=0:
            infl_res[j][1] += 1
            infl_res[j][2] += infl[j]
            if t['flipped_idx'][0]:
                infl_res[j][10] += 1
            else:
                infl_res[j][11] += 1
        else:
            infl_res[j][3] += 1
            infl_res[j][4] += infl[j]
            if t['flipped_idx'][0]:
                infl_res[j][12] += 1
            else:
                infl_res[j][13] += 1
   
np.savez(BASE_DIR + 'inception_influence_result', result = infl_res)

wb = Workbook()
ws = wb.active
for title_idx in range(18):
    ws.cell(row = 1, column = 1, value = 'index')
    ws.cell(row = 1, column = title_idx + 2, value = index_titles[title_idx])
for i in range(1800):
    ws.cell(row = i+2, column = 1, value = i)
    for j in range(18):
        ws.cell(row = i+2, column = j+2, value = infl_res[i][j])

wb.save(BASE_DIR + "inception_influence_result_dogfish_top5.xlsx")

#%%
# 把top5中训练点的标签，和测试点是否一致给区别开来。
# =============================================================================
# index:index
# 0:label
# 1:postive influence count
# 2:postive influence sum
# 3:negative influence count
# 4:negative influence sum
# 5:top5 count
# 6:top5 influence sum
# 7:last5 count
# 8:last5 influence sum
# 9:influence sum
# 10:postive influence count same cata
# 11:postive influence count diff cata
# 12:negative influence count same cata
# 13:negative influence count diff cata
# 14:top5 count same cata
# 15:top5 count diff cata
# 16:last5 count same cata
# 17:last5 count diff cata
# =============================================================================
# index不用单独一列存，加了是否和测试点标签一致之后，执行要7分钟了。。
infl_res = np.zeros((2000, 18))
for i in range(600):
    t = np.load(BASE_DIR + RESULT_STR + '_%s.npz' % i)
    infl = t['inception_predicted_loss_diffs']
    for idx in np.argsort(infl)[-5:]:
        infl_res[idx][5] += 1
        infl_res[idx][6] += infl[idx]
        if t['flipped_idx'][0]:
            infl_res[idx][14] += 1
        else:
            infl_res[idx][15] += 1
    for idx in np.argsort(infl)[:5]:
        infl_res[idx][7] += 1
        infl_res[idx][8] += infl[idx]
        if t['flipped_idx'][0]:
            infl_res[idx][16] += 1
        else:
            infl_res[idx][17] += 1
    for j in range(2000):
        infl_res[j][0] = Y_train[j]
        infl_res[j][9] += infl[j]
        if infl[j]>=0:
            infl_res[j][1] += 1
            infl_res[j][2] += infl[j]
            if t['flipped_idx'][0]:
                infl_res[j][10] += 1
            else:
                infl_res[j][11] += 1
        else:
            infl_res[j][3] += 1
            infl_res[j][4] += infl[j]
            if t['flipped_idx'][0]:
                infl_res[j][12] += 1
            else:
                infl_res[j][13] += 1

        
np.savez(BASE_DIR + 'inception_influence_result_carair_top5.npz', result = infl_res)

index_titles = {0:'label',
                1:'postive influence count',
                2:'postive influence sum',
                3:'negative influence count',
                4:'negative influence sum',
                5:'top5 count',
                6:'top5 influence sum',
                7:'last5 count',
                8:'last5 influence sum',
                9:'influence sum',
                10:'postive influence count same cata',
                11:'postive influence count diff cata',
                12:'negative influence count same cata',
                13:'negative influence count diff cata',
                14:'top5 count same cata',
                15:'top5 count diff cata',
                16:'last5 count same cata',
                17:'last5 count diff cata'}
wb = Workbook()
ws = wb.active
for title_idx in range(18):
    ws.cell(row = 1, column = 1, value = 'index')
    ws.cell(row = 1, column = title_idx + 2, value = index_titles[title_idx])
for i in range(2000):
    ws.cell(row = i+2, column = 1, value = i)
    for j in range(18):
        ws.cell(row = i+2, column = j+2, value = infl_res[i][j])

wb.save(BASE_DIR + "inception_influence_result_carair_top5.xlsx")










